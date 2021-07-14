import openpyxl
import smtplib
import pandas as pd

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from bs4 import BeautifulSoup

# SMTP information
email_user = "########" # Change this and add your SMTP user info
email_pass = "########" # Change this and add your SMTP pass info

# Get Excel spreadsheet
filename = ".xslx file location here" # Change this to where your Excel file location is

# Get active sheet
wb = openpyxl.load_workbook(filename)
ws = wb.active

class Items:
  def __init__(self, item_code, item_descript, order_qty):
    self.item_code = item_code.split(',')
    self.item_descript = item_descript.split(';')
    self.order_qty = order_qty.split(',')
    
    self.item_list = []
    
    for i in range(0, len(self.item_code)):
      temp = '''
        <tr class="item">
          <td class="add-info"><p>'''+ self.item_code[i] +'''</p></td>
          <td class="td-wide" colspan="2"><p>'''+ self.item_descript[i] +'''</p></td>
          <td class="add-info align-center"><p>'''+ self.order_qty[i] +'''</p></td>
        </tr>
      '''
      self.item_list.append(temp)
  
def build_email(ordnum, ordate, ship_name, ship_addr1, ship_addr2, ship_addr3, ship_city, ship_state, ship_zipcode, c):
  # Get Items
  code = descript = order = []
  
  itemtemp = Items(get_data(code, 1)[c], get_data(descript, 2)[c], get_data(order, 3)[c])
  
  # Create email body
  html = '''\
    <!DOCTYPE HTML PUBLIC "-//W3C//DTD HTML 4.01 Transitional//EN" "http://www.w3.org/TR/html4/loose.dtd">
      <html xmlns="http://www.w3.org/1999/xhtml">
        <head>
          <meta http-equiv="Content-Type" content="text/html; charset=UTF-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <style type="text/css">
            .container {
              max-width: 920px;
              margin: 0 auto;
              padding: 40px;
              width: 100%;
              border: 1px solid #eee;
              box-shadow: 0 0 10px rgba(0, 0, 0, .15);
              line-height: 22px;
            }
            
            table {
              width: 100%;
              text-align: left;
            }
            
            .title {
              font-weight: bold;
            }
            
            .heading {
              background: #e4e4e4;
              font-weight: bold;
            }
            
            .heading td {
              padding: 0 8px;
            }
            
            .align-center {
              text-align: center;
            }
            
            .status {
              max-width: 980px;
              margin: 0 auto;
              width: 100%;
            }
          </style>
        </head>
        <body>
          <h1 style="text-align: center;">Order Status</h1>
          <div class="status">
            <p>Hello '''+ ship_name +''',</p>
            <p>Thank you for your order! Listed below shows all of your order details.</p>
            <p>Regards,<br>Grace</p>
          </div>
          <div class="container">
            <table class="info-table" cellpadding="0" cellspacing="0">
              <tr class="head-info">
                <td class="logo">Logo Here</td>
                <td class="td-wide" colspan="2"></td>
                <td><p class="align-right">Order Date: '''+ ordate +'''</p></td>
              </tr>
              <tr class="company">
                <td><p class="title">Blue Rock Foundation<br>2345 Northlake Ave<br>Duluth, GA 30026<br>(678)-123-4567</p></td>
                <td class="td-wide" colspan="2"></td>
              </tr>
              <tr class="heading">
                <td><p>Ship To Address:</p></td>
                <td class="td-wide" colspan="3"><p></p></td>
              </tr>
              <tr>
                <td class="add-info align-right"><p class="ship-name">'''+ ship_name +'''</p><p class="ship-add1">'''+ ship_addr1 +'''</p><p class="ship-add2">'''+ ship_addr2 +'''</p><p class="ship-add3">'''+ ship_addr3 +'''</p><p class="ship-statezip">'''+ ship_city +''', '''+ ship_state +''' '''+ ship_zipcode +'''</p></td>
              </tr>
              <tr>
                <td><p>Order #</p></td>
                <td class="td-wide" colspan="3"><p></p></td>
              </tr>
              <tr>
                <td class="add-info"><p class="order-num">'''+ ordnum +'''</p></td>
              </tr>
              <tr class="heading">
                <td><p>Item Name</p></td>
                <td class="td-wide" colspan="2"><p>Description</p></td>
                <td><p class="align-center">Order Qty</p></td>
              </tr>
              ''' + ''.join(itemtemp.item_list) +'''
            </table>
          </div>
        </body>
      </html>
  '''
  
  # Using BeautifulSoup to get rid of None type where cell values are blank
  soup = BeautifulSoup(html, 'html.parser')
  if ship_addr1 == 'None':
    soup.find('p', {'class' : 'ship-add1'}).decompose()
  if ship_addr2 == 'None':
    soup.find('p', {'class' : 'ship-add2'}).decompose()
  if ship_addr3 == 'None':
    soup.find('p', {'class' : 'ship-add3'}).decompose()
  
  return soup.prettify()
  

def send_email(user, password, sent_from, to, msg):
  # Enter SMTP credentials
  server = smtplib.SMTP('smtp.gmail.com', 587) # Change this and enter SMTP (gmail, outlook, etc.) Settings and SMTP port number
  server.ehlo()
  server.starttls()
  server.ehlo()
  server.login(user, password)
  server.sendmail(sent_from, to, msg.as_string())
  server.quit()

# Get column name and column number  
def get_data(colname, colnum):
  colname = []
  count = 1
  
  # Go through each row in .xlsx file
  # If there are headers, then skip first row
  for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    #print(f"Row {count}: {row[colnum].value}")
    colname.append(str(row[colnum].value))
    count += 1
  
  return colname

#print(get_data(onum, 0))

count = 0

# Create empty list for each column name
onum = orderdate = shipname = shipaddr1 = shipaddr2 = shipaddr3 = shipcity = shipstate = shipzip = emails = []

# Find column number where emails are listed on the file and loop through them
for to in get_data(emails, 12):
  msg = MIMEMultipart('alternative')
  msg['To'] = get_data(emails, 12)[count]
  msg['From'] = "Blue Rock Foundation <bluerock@gmail.com>"
  msg['Subject'] = "Order Number " + get_data(onum, 0) + " Information"
  body = MIMEText(build_email(get_data(onum, 0)[count], get_data(orderdate, 4)[count], get_data(shipname, 5)[count], get_data(shipaddr1, 6)[count], get_data(shipaddr2, 7)[count], get_data(shipaddr3, 8)[count], get_data(shipcity, 9)[count], get_data(shipstate, 10)[count], get_data(shipzip, 11)[count], count), html)
  msg.attach(body)
  
  try:
    send_email(email_user, email_pass, msg['From'], msg['To'], msg)
    print("Email successfully sent to: ", get_data(emails, 12)[count])
    count += 1
  except Exception as e:
    print(e)
    print("Email failed to send to: ", get_data(emails, 12)[count])
